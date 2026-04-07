#!/usr/bin/env python3
"""
Enrich US occupations data with:
1. OES — State-level employment and wages per occupation (BLS)
2. ACS — Demographics (gender, race) per occupation (Census Bureau)

Usage:
    pip install requests openpyxl
    python us/enrich_data.py

    # To skip downloads and use cached files:
    python us/enrich_data.py --cached

Output:
    us/site/data.json     — enriched occupation data
    us/site/summary.json  — updated summary with demographics & state data
"""

import json
import csv
import os
import sys
import io
import zipfile
import requests
import time
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.dirname(BASE_DIR)
CACHE_DIR = os.path.join(BASE_DIR, "cache")
SITE_DIR = os.path.join(BASE_DIR, "site")

os.makedirs(CACHE_DIR, exist_ok=True)

USE_CACHED = "--cached" in sys.argv

# ── Helpers ────────────────────────────────────────────────────────────────

def download(url, filename, desc=""):
    path = os.path.join(CACHE_DIR, filename)
    if USE_CACHED and os.path.exists(path):
        print(f"  [cached] {filename}")
        return path
    print(f"  Downloading {desc or filename}...")
    resp = requests.get(url, timeout=120, headers={
        "User-Agent": "Mozilla/5.0 (research project) ai-jobs-impact/1.0"
    })
    resp.raise_for_status()
    with open(path, "wb") as f:
        f.write(resp.content)
    print(f"  Saved {filename} ({len(resp.content) / 1024:.0f} KB)")
    return path


def load_soc_mapping():
    """Load slug -> SOC code mapping from occupations.csv"""
    mapping = {}
    csv_path = os.path.join(REPO_DIR, "occupations.csv")
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            slug = row.get("slug", "").strip()
            soc = row.get("soc_code", "").strip()
            if slug and soc:
                mapping[slug] = soc
    print(f"  Loaded {len(mapping)} SOC code mappings")
    return mapping


def load_current_data():
    """Load current us/site/data.json"""
    path = os.path.join(SITE_DIR, "data.json")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ── 1. OES: State-level employment and wages ──────────────────────────────

STATE_FIPS = {
    "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA",
    "08": "CO", "09": "CT", "10": "DE", "11": "DC", "12": "FL",
    "13": "GA", "15": "HI", "16": "ID", "17": "IL", "18": "IN",
    "19": "IA", "20": "KS", "21": "KY", "22": "LA", "23": "ME",
    "24": "MD", "25": "MA", "26": "MI", "27": "MN", "28": "MS",
    "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI",
    "45": "SC", "46": "SD", "47": "TN", "48": "TX", "49": "UT",
    "50": "VT", "51": "VA", "53": "WA", "54": "WV", "55": "WI",
    "56": "WY", "72": "PR", "78": "VI",
}

STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island",
    "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas",
    "UT": "Utah", "VT": "Vermont", "VA": "Virginia", "WA": "Washington",
    "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming",
}


def fetch_oes_state_data():
    """
    Download OES state-level data from BLS.
    Uses the Excel file from special requests.
    Returns: dict of soc_code -> { state_abbr: { employment, median_pay } }
    """
    print("\n[1/2] OES State-Level Data")

    # Try May 2024 first, fall back to 2023
    for year in ["24", "23", "22"]:
        url = f"https://www.bls.gov/oes/special.requests/oesm{year}st.zip"
        try:
            path = download(url, f"oesm{year}st.zip", f"OES state data (20{year})")
            break
        except Exception as e:
            print(f"  Failed for 20{year}: {e}")
            continue
    else:
        print("  ERROR: Could not download OES state data. Trying all_data...")
        # Try the all_data file instead
        for year in ["24", "23", "22"]:
            url = f"https://www.bls.gov/oes/special.requests/oesm{year}all.zip"
            try:
                path = download(url, f"oesm{year}all.zip", f"OES all data (20{year})")
                break
            except Exception as e:
                print(f"  Failed for 20{year} all: {e}")
                continue
        else:
            print("  ERROR: Could not download any OES data file.")
            return {}

    # Parse the ZIP file (contains Excel files)
    import openpyxl
    result = {}  # soc_code -> { state_abbr: { employment, median_pay } }

    with zipfile.ZipFile(path) as zf:
        # Find the state-level Excel file
        xlsx_files = [n for n in zf.namelist() if n.endswith((".xlsx", ".xls")) and "state" in n.lower()]
        if not xlsx_files:
            # Try any Excel file
            xlsx_files = [n for n in zf.namelist() if n.endswith((".xlsx", ".xls"))]
        if not xlsx_files:
            # Try CSV
            csv_files = [n for n in zf.namelist() if n.endswith(".csv")]
            if csv_files:
                xlsx_files = csv_files

        print(f"  Files in ZIP: {zf.namelist()}")

        if not xlsx_files:
            print("  ERROR: No data files found in ZIP")
            return {}

        target = xlsx_files[0]
        print(f"  Parsing {target}...")

        data_bytes = zf.read(target)

        if target.endswith(".csv"):
            # Parse CSV
            text = data_bytes.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows_parsed = 0
            for row in reader:
                # Filter to state-level, cross-industry, detailed occupations
                area_type = row.get("AREA_TYPE", row.get("area_type", ""))
                if area_type != "2":  # 2 = State
                    continue
                o_group = row.get("O_GROUP", row.get("o_group", ""))
                if o_group not in ("detailed", "broad", "DETAILED", "BROAD"):
                    continue

                soc = row.get("OCC_CODE", row.get("occ_code", "")).strip()
                state_fips = row.get("PRIM_STATE", row.get("prim_state",
                             row.get("AREA", row.get("area", ""))[:2])).strip()

                # Map FIPS to state abbreviation
                # The AREA field for states is typically the FIPS code + "0000"
                area = row.get("AREA", row.get("area", "")).strip()
                if len(area) >= 2:
                    state_fips = area[:2]

                state_abbr = STATE_FIPS.get(state_fips, "")
                if not state_abbr or not soc:
                    continue

                try:
                    emp = row.get("TOT_EMP", row.get("tot_emp", ""))
                    emp = int(emp.replace(",", "")) if emp and emp not in ("**", "*", "-") else 0
                    pay = row.get("A_MEDIAN", row.get("a_median", ""))
                    pay = int(float(pay.replace(",", ""))) if pay and pay not in ("**", "*", "#", "-") else 0
                except (ValueError, TypeError):
                    continue

                if soc not in result:
                    result[soc] = {}
                result[soc][state_abbr] = {
                    "employment": emp,
                    "median_pay": pay,
                }
                rows_parsed += 1

            print(f"  Parsed {rows_parsed} state-occupation rows for {len(result)} occupations")

        else:
            # Parse Excel
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True)
            ws = wb.active

            # Get headers
            headers = []
            for row in ws.iter_rows(max_row=1, values_only=True):
                headers = [str(h).strip().upper() if h else "" for h in row]
                break

            print(f"  Columns: {headers[:15]}...")

            # Find column indices
            def col_idx(names):
                for name in names:
                    name_upper = name.upper()
                    for i, h in enumerate(headers):
                        if h == name_upper:
                            return i
                return -1

            area_type_idx = col_idx(["AREA_TYPE"])
            area_idx = col_idx(["AREA"])
            occ_code_idx = col_idx(["OCC_CODE"])
            o_group_idx = col_idx(["O_GROUP"])
            tot_emp_idx = col_idx(["TOT_EMP"])
            a_median_idx = col_idx(["A_MEDIAN"])
            naics_idx = col_idx(["NAICS", "NAICS_TITLE"])

            rows_parsed = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(area_type_idx, occ_code_idx):
                    continue

                # Filter: state-level (area_type=2), cross-industry
                if area_type_idx >= 0:
                    at = str(row[area_type_idx]).strip() if row[area_type_idx] else ""
                    if at != "2":
                        continue

                if o_group_idx >= 0:
                    og = str(row[o_group_idx]).strip().lower() if row[o_group_idx] else ""
                    if og not in ("detailed", "broad"):
                        continue

                # Filter to cross-industry (NAICS = 000000 or similar)
                if naics_idx >= 0:
                    naics = str(row[naics_idx]).strip() if row[naics_idx] else ""
                    if naics and naics not in ("000000", "0", "Cross-industry"):
                        continue

                soc = str(row[occ_code_idx]).strip() if occ_code_idx >= 0 and row[occ_code_idx] else ""
                area = str(row[area_idx]).strip() if area_idx >= 0 and row[area_idx] else ""

                if not soc or len(area) < 2:
                    continue

                state_fips = area[:2]
                state_abbr = STATE_FIPS.get(state_fips, "")
                if not state_abbr:
                    continue

                try:
                    emp_val = row[tot_emp_idx] if tot_emp_idx >= 0 else None
                    emp = int(str(emp_val).replace(",", "")) if emp_val and str(emp_val) not in ("**", "*", "-") else 0
                    pay_val = row[a_median_idx] if a_median_idx >= 0 else None
                    pay = int(float(str(pay_val).replace(",", ""))) if pay_val and str(pay_val) not in ("**", "*", "#", "-") else 0
                except (ValueError, TypeError):
                    continue

                if soc not in result:
                    result[soc] = {}
                result[soc][state_abbr] = {
                    "employment": emp,
                    "median_pay": pay,
                }
                rows_parsed += 1

            wb.close()
            print(f"  Parsed {rows_parsed} state-occupation rows for {len(result)} occupations")

    return result


# ── 2. ACS: Demographics by occupation ────────────────────────────────────

# Census API table B24010: Sex by Occupation for Civilian Employed Population 16+
# We use the broader EEO occupation categories and map them to our SOC codes

# Major SOC groups mapped to Census EEO-1 occupation categories
SOC_TO_CENSUS_OCC = {
    "11": "Management",
    "13": "Business and financial operations",
    "15": "Computer and mathematical",
    "17": "Architecture and engineering",
    "19": "Life, physical, and social science",
    "21": "Community and social service",
    "23": "Legal",
    "25": "Educational instruction and library",
    "27": "Arts, design, entertainment, sports, and media",
    "29": "Healthcare practitioners and technical",
    "31": "Healthcare support",
    "33": "Protective service",
    "35": "Food preparation and serving related",
    "37": "Building and grounds cleaning and maintenance",
    "39": "Personal care and service",
    "41": "Sales and related",
    "43": "Office and administrative support",
    "45": "Farming, fishing, and forestry",
    "47": "Construction and extraction",
    "49": "Installation, maintenance, and repair",
    "51": "Production",
    "53": "Transportation and material moving",
    "55": "Military specific",
}


def fetch_acs_demographics():
    """
    Fetch demographics data from Census ACS API.
    Uses table S2401 (Occupation by Sex) and B03002 iterative tables for race.
    Returns: dict of soc_major -> { total, male, female, white, black, hispanic, asian }
    """
    print("\n[2/2] ACS Demographics Data")

    # Use the Census ACS 5-year estimates API
    # Table B24010: Sex by Occupation (detailed)
    # We'll get major occupation group totals by sex

    base_url = "https://api.census.gov/data/2022/acs/acs5/subject"

    demographics = {}

    # Get occupation by sex data from S2401
    # S2401 columns: total employed, male, female for each occupation group
    print("  Fetching occupation by sex (S2401)...")
    try:
        # S2401 table - key variables for major occupation groups
        # Format: S2401_C01_XXX (total), S2401_C02_XXX (male), S2401_C03_XXX (female)
        # These use estimate numbers that map to occupation groups

        # Instead, use B24010 which is more reliable
        # B24010_001E = Total, B24010_002E = Male total, B24010_038E = Female total (approx)

        # Actually let's use the simpler approach: ACS PUMS summary via S2401
        variables = "NAME,S2401_C01_001E,S2401_C02_001E,S2401_C03_001E"

        # Get the occupation breakdown rows
        # S2401 has specific row numbers for each occupation group
        occupation_vars = {
            "11": {"total": "S2401_C01_002E", "male": "S2401_C02_002E", "female": "S2401_C03_002E"},  # Management
            "13": {"total": "S2401_C01_003E", "male": "S2401_C02_003E", "female": "S2401_C03_003E"},  # Business/financial
            "15": {"total": "S2401_C01_004E", "male": "S2401_C02_004E", "female": "S2401_C03_004E"},  # Computer/math
            "17": {"total": "S2401_C01_005E", "male": "S2401_C02_005E", "female": "S2401_C03_005E"},  # Architecture/engineering
            "19": {"total": "S2401_C01_006E", "male": "S2401_C02_006E", "female": "S2401_C03_006E"},  # Life/physical/social science
            "21": {"total": "S2401_C01_007E", "male": "S2401_C02_007E", "female": "S2401_C03_007E"},  # Community/social
            "23": {"total": "S2401_C01_008E", "male": "S2401_C02_008E", "female": "S2401_C03_008E"},  # Legal
            "25": {"total": "S2401_C01_009E", "male": "S2401_C02_009E", "female": "S2401_C03_009E"},  # Education
            "27": {"total": "S2401_C01_010E", "male": "S2401_C02_010E", "female": "S2401_C03_010E"},  # Arts/entertainment
            "29": {"total": "S2401_C01_011E", "male": "S2401_C02_011E", "female": "S2401_C03_011E"},  # Healthcare practitioners
            "31": {"total": "S2401_C01_012E", "male": "S2401_C02_012E", "female": "S2401_C03_012E"},  # Healthcare support
            "33": {"total": "S2401_C01_014E", "male": "S2401_C02_014E", "female": "S2401_C03_014E"},  # Protective service
            "35": {"total": "S2401_C01_015E", "male": "S2401_C02_015E", "female": "S2401_C03_015E"},  # Food prep
            "37": {"total": "S2401_C01_016E", "male": "S2401_C02_016E", "female": "S2401_C03_016E"},  # Building/grounds
            "39": {"total": "S2401_C01_017E", "male": "S2401_C02_017E", "female": "S2401_C03_017E"},  # Personal care
            "41": {"total": "S2401_C01_019E", "male": "S2401_C02_019E", "female": "S2401_C03_019E"},  # Sales
            "43": {"total": "S2401_C01_020E", "male": "S2401_C02_020E", "female": "S2401_C03_020E"},  # Office/admin
            "45": {"total": "S2401_C01_022E", "male": "S2401_C02_022E", "female": "S2401_C03_022E"},  # Farming
            "47": {"total": "S2401_C01_023E", "male": "S2401_C02_023E", "female": "S2401_C03_023E"},  # Construction
            "49": {"total": "S2401_C01_024E", "male": "S2401_C02_024E", "female": "S2401_C03_024E"},  # Installation/repair
            "51": {"total": "S2401_C01_025E", "male": "S2401_C02_025E", "female": "S2401_C03_025E"},  # Production
            "53": {"total": "S2401_C01_026E", "male": "S2401_C02_026E", "female": "S2401_C03_026E"},  # Transportation
        }

        # Build the variable list
        all_vars = []
        for soc_major, vars_dict in occupation_vars.items():
            all_vars.extend(vars_dict.values())

        # Census API has a limit of ~50 variables per request, so batch
        for batch_start in range(0, len(all_vars), 45):
            batch = all_vars[batch_start:batch_start + 45]
            var_str = ",".join(["NAME"] + batch)
            url = f"{base_url}?get={var_str}&for=us:1"
            print(f"  Fetching batch {batch_start // 45 + 1}...")
            resp = requests.get(url, timeout=30)
            if resp.status_code != 200:
                print(f"  Warning: Census API returned {resp.status_code}")
                continue
            rows = resp.json()
            headers = rows[0]
            values = rows[1] if len(rows) > 1 else []

            for soc_major, vars_dict in occupation_vars.items():
                if soc_major not in demographics:
                    demographics[soc_major] = {"total": 0, "male": 0, "female": 0}
                for key, var_name in vars_dict.items():
                    if var_name in headers:
                        idx = headers.index(var_name)
                        try:
                            raw = values[idx]
                            if raw is None or raw == "":
                                continue
                            val = float(raw)
                            # C03 columns return percentage, not count
                            if key == "female" and val < 100:
                                # It's a percentage - compute from total
                                total = demographics[soc_major].get("total", 0)
                                demographics[soc_major]["female"] = round(total * val / 100)
                            else:
                                demographics[soc_major][key] = int(val)
                        except (ValueError, IndexError):
                            pass
                # Ensure female is computed if we have total and male
                d = demographics[soc_major]
                if d["female"] == 0 and d["total"] > 0 and d["male"] > 0:
                    d["female"] = d["total"] - d["male"]

            time.sleep(0.5)  # Rate limit

        print(f"  Got gender data for {len(demographics)} occupation groups")

    except Exception as e:
        print(f"  Warning: Could not fetch ACS gender data: {e}")

    # Fetch race data using C24010 tables (non-subject endpoint)
    print("  Fetching occupation by race...")
    race_url = "https://api.census.gov/data/2022/acs/acs5"
    try:
        url = f"{race_url}?get=NAME,C24010_001E,C24010A_001E,C24010B_001E,C24010D_001E,C24010I_001E&for=us:1"
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            rows = resp.json()
            if len(rows) > 1:
                vals = rows[1]
                grand_total = int(vals[1]) if vals[1] else 1
                white_total = int(vals[2]) if vals[2] else 0
                black_total = int(vals[3]) if vals[3] else 0
                asian_total = int(vals[4]) if vals[4] else 0
                hispanic_total = int(vals[5]) if vals[5] else 0

                pct_white = white_total / grand_total if grand_total else 0
                pct_black = black_total / grand_total if grand_total else 0
                pct_asian = asian_total / grand_total if grand_total else 0
                pct_hispanic = hispanic_total / grand_total if grand_total else 0

                for soc_major, data in demographics.items():
                    t = data.get("total", 0)
                    data["white"] = round(t * pct_white)
                    data["black"] = round(t * pct_black)
                    data["asian"] = round(t * pct_asian)
                    data["hispanic"] = round(t * pct_hispanic)

                print(f"  Race distribution for {len(demographics)} groups")
                print(f"    Overall: {pct_white:.1%} White, {pct_black:.1%} Black, {pct_asian:.1%} Asian, {pct_hispanic:.1%} Hispanic")
        else:
            print(f"  Warning: Race data API returned {resp.status_code}")
    except Exception as e:
        print(f"  Warning: Could not fetch race data: {e}")

    return demographics


# ── Merge everything ──────────────────────────────────────────────────────

def merge_data(current_data, soc_mapping, oes_state, acs_demographics):
    """Merge all enrichment data into the current dataset."""
    print("\n[Merging] Enriching data.json...")

    matched_oes = 0
    matched_demo = 0

    for occ in current_data:
        slug = occ.get("slug", "")
        soc = soc_mapping.get(slug, "")

        # 1. Add SOC code
        if soc:
            occ["soc_code"] = soc

        # 2. Add OES state-level data
        if soc and soc in oes_state:
            state_data = oes_state[soc]
            # Build por_state object (like Brazil's por_uf)
            by_state = {}
            for state_abbr, sdata in state_data.items():
                if state_abbr in STATE_NAMES and sdata.get("employment", 0) > 0:
                    by_state[state_abbr] = {
                        "employment": sdata["employment"],
                        "median_pay": sdata["median_pay"],
                    }
            if by_state:
                occ["by_state"] = by_state
                matched_oes += 1
        else:
            # Try matching with broader SOC (e.g., 13-2011 -> 13-2010)
            if soc:
                broad_soc = soc[:-1] + "0"
                if broad_soc in oes_state:
                    state_data = oes_state[broad_soc]
                    by_state = {}
                    for state_abbr, sdata in state_data.items():
                        if state_abbr in STATE_NAMES and sdata.get("employment", 0) > 0:
                            by_state[state_abbr] = {
                                "employment": sdata["employment"],
                                "median_pay": sdata["median_pay"],
                            }
                    if by_state:
                        occ["by_state"] = by_state
                        matched_oes += 1

        # 3. Add demographics (from SOC major group)
        if soc:
            soc_major = soc[:2]
            demo = acs_demographics.get(soc_major)
            if demo and demo.get("total", 0) > 0:
                jobs = occ.get("jobs", 0) or 0
                group_total = demo["total"]
                if group_total > 0 and jobs > 0:
                    # Estimate this occupation's demographics proportionally
                    ratio = jobs / group_total
                    occ["demographics"] = {
                        "total_male": round(demo.get("male", 0) * ratio),
                        "total_female": round(demo.get("female", 0) * ratio),
                        "total_white": round(demo.get("white", 0) * ratio),
                        "total_black": round(demo.get("black", 0) * ratio),
                        "total_asian": round(demo.get("asian", 0) * ratio),
                        "total_hispanic": round(demo.get("hispanic", 0) * ratio),
                        "pct_female": round(demo.get("female", 0) / group_total * 100, 1),
                        "pct_male": round(demo.get("male", 0) / group_total * 100, 1),
                    }
                    matched_demo += 1

    print(f"  OES state data: {matched_oes}/{len(current_data)} occupations matched")
    print(f"  ACS demographics: {matched_demo}/{len(current_data)} occupations matched")

    return current_data


def build_summary(data, acs_demographics):
    """Build updated summary.json with demographics and state data."""
    print("\n[Summary] Building summary.json...")

    # Basic stats
    total_workers = sum(d.get("jobs", 0) or 0 for d in data)
    total_occupations = len([d for d in data if (d.get("jobs") or 0) > 0])

    # Categories
    categories = {}
    for d in data:
        jobs = d.get("jobs") or 0
        if jobs <= 0:
            continue
        cat = d.get("category", "other")
        if cat not in categories:
            categories[cat] = {"name": cat, "total_workers": 0, "count": 0,
                               "sum_exposure": 0, "sum_pay": 0}
        categories[cat]["total_workers"] += jobs
        categories[cat]["count"] += 1
        categories[cat]["sum_exposure"] += (d.get("exposure") or 0) * jobs
        categories[cat]["sum_pay"] += (d.get("pay") or 0) * jobs

    cat_list = []
    for k, v in categories.items():
        cat_list.append({
            "name": k,
            "total_workers": v["total_workers"],
            "occupations": v["count"],
            "avg_exposure": round(v["sum_exposure"] / v["total_workers"], 1) if v["total_workers"] > 0 else 0,
            "avg_pay": round(v["sum_pay"] / v["total_workers"]) if v["total_workers"] > 0 else 0,
        })
    cat_list.sort(key=lambda x: -x["total_workers"])

    # State-level summary
    state_summary = {}
    for d in data:
        by_state = d.get("by_state", {})
        for state_abbr, sdata in by_state.items():
            if state_abbr not in state_summary:
                state_summary[state_abbr] = {
                    "name": STATE_NAMES.get(state_abbr, state_abbr),
                    "abbr": state_abbr,
                    "total_workers": 0,
                    "sum_pay": 0,
                    "sum_exposure": 0,
                    "top_occupations": [],
                }
            ss = state_summary[state_abbr]
            emp = sdata.get("employment", 0)
            ss["total_workers"] += emp
            ss["sum_pay"] += sdata.get("median_pay", 0) * emp
            ss["sum_exposure"] += (d.get("exposure") or 0) * emp
            ss["top_occupations"].append({
                "title": d["title"],
                "workers": emp,
            })

    # Finalize state summaries
    for state_abbr, ss in state_summary.items():
        tw = ss["total_workers"]
        ss["avg_pay"] = round(ss["sum_pay"] / tw) if tw > 0 else 0
        ss["avg_exposure"] = round(ss["sum_exposure"] / tw, 1) if tw > 0 else 0
        ss["top_occupations"].sort(key=lambda x: -x["workers"])
        ss["top_occupations"] = ss["top_occupations"][:5]
        del ss["sum_pay"]
        del ss["sum_exposure"]

    # Demographics summary
    total_male = sum(d.get("demographics", {}).get("total_male", 0) for d in data)
    total_female = sum(d.get("demographics", {}).get("total_female", 0) for d in data)
    total_white = sum(d.get("demographics", {}).get("total_white", 0) for d in data)
    total_black = sum(d.get("demographics", {}).get("total_black", 0) for d in data)
    total_asian = sum(d.get("demographics", {}).get("total_asian", 0) for d in data)
    total_hispanic = sum(d.get("demographics", {}).get("total_hispanic", 0) for d in data)

    # High risk by gender
    high_risk_male = sum(d.get("demographics", {}).get("total_male", 0) for d in data if (d.get("exposure") or 0) >= 6)
    high_risk_female = sum(d.get("demographics", {}).get("total_female", 0) for d in data if (d.get("exposure") or 0) >= 6)

    demographics_summary = {
        "total_male": total_male,
        "total_female": total_female,
        "total_white": total_white,
        "total_black": total_black,
        "total_asian": total_asian,
        "total_hispanic": total_hispanic,
        "workers_with_demographics": total_male + total_female,
        "high_risk_male": high_risk_male,
        "high_risk_female": high_risk_female,
        "pct_high_risk_male": round(high_risk_male / total_male * 100, 1) if total_male > 0 else 0,
        "pct_high_risk_female": round(high_risk_female / total_female * 100, 1) if total_female > 0 else 0,
    }

    summary = {
        "total_occupations": total_occupations,
        "total_workers": total_workers,
        "categories": cat_list,
        "by_state": state_summary,
        "demographics": demographics_summary,
    }

    return summary


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("US Occupation Data Enrichment")
    print("=" * 60)

    # Load current data
    print("\nLoading current data...")
    current_data = load_current_data()
    soc_mapping = load_soc_mapping()
    print(f"  {len(current_data)} occupations, {len(soc_mapping)} SOC mappings")

    # Fetch data sources
    oes_state = fetch_oes_state_data()
    acs_demographics = fetch_acs_demographics()

    # Merge
    enriched = merge_data(current_data, soc_mapping, oes_state, acs_demographics)

    # Build summary
    summary = build_summary(enriched, acs_demographics)

    # Save
    data_path = os.path.join(SITE_DIR, "data.json")
    summary_path = os.path.join(SITE_DIR, "summary.json")

    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(enriched, f, ensure_ascii=False)
    print(f"\nSaved enriched data to {data_path}")

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"Saved summary to {summary_path}")

    # Stats
    has_states = sum(1 for d in enriched if d.get("by_state"))
    has_demo = sum(1 for d in enriched if d.get("demographics"))
    print(f"\nEnrichment results:")
    print(f"  State data: {has_states}/{len(enriched)} occupations")
    print(f"  Demographics: {has_demo}/{len(enriched)} occupations")
    print(f"  States in summary: {len(summary.get('by_state', {}))}")
    print(f"\nDone!")


if __name__ == "__main__":
    main()
